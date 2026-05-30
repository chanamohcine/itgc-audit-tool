# ============================================================
# UTILITAIRE : Mise en forme professionnelle du rapport Excel
# ============================================================

from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ----------------------------------------------------------
# PALETTE DE COULEURS
# ----------------------------------------------------------
ROUGE        = "C0392B"   # Critique
ORANGE       = "E67E22"   # Élevé
JAUNE        = "F1C40F"   # Moyen
VERT         = "27AE60"   # OK
BLEU_HEADER  = "1F3864"   # En-têtes
GRIS_CLAIR   = "F2F2F2"   # Lignes alternées
BLANC        = "FFFFFF"

def couleur_criticite(valeur):
    """Retourne la couleur selon le niveau de criticité."""
    if "CRITIQUE" in str(valeur):
        return ROUGE
    elif "ÉLEVÉ" in str(valeur):
        return ORANGE
    elif "MOYEN" in str(valeur):
        return JAUNE
    return VERT

def formater_onglet(ws):
    """Applique la mise en forme à un onglet."""

    # -- En-têtes : fond bleu foncé, texte blanc, gras --
    header_fill = PatternFill("solid", fgColor=BLEU_HEADER)
    header_font = Font(color=BLANC, bold=True, size=10,
                       name="Calibri")
    header_align = Alignment(horizontal="center",
                              vertical="center", wrap_text=True)

    # Bordure fine pour toutes les cellules
    bordure = Border(
        left   = Side(style="thin", color="CCCCCC"),
        right  = Side(style="thin", color="CCCCCC"),
        top    = Side(style="thin", color="CCCCCC"),
        bottom = Side(style="thin", color="CCCCCC")
    )

    nb_colonnes = ws.max_column
    nb_lignes   = ws.max_row

    # Appliquer le style sur la ligne d'en-tête (ligne 1)
    for col in range(1, nb_colonnes + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = bordure

    # Figer la première ligne (en-têtes toujours visibles)
    ws.freeze_panes = "A2"

    # -- Lignes de données --
    for row in range(2, nb_lignes + 1):
        # Couleur de fond alternée (zébrage)
        bg_color = BLANC if row % 2 == 0 else GRIS_CLAIR
        fill_ligne = PatternFill("solid", fgColor=bg_color)

        for col in range(1, nb_colonnes + 1):
            cell = ws.cell(row=row, column=col)
            cell.border    = bordure
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=True)
            cell.font      = Font(size=9, name="Calibri")

            # Colorer la cellule "criticite" selon son niveau
            if ws.cell(row=1, column=col).value == "criticite":
                couleur = couleur_criticite(cell.value)
                cell.fill = PatternFill("solid", fgColor=couleur)
                cell.font = Font(color=BLANC, bold=True,
                                 size=9, name="Calibri")
            else:
                cell.fill = fill_ligne

    # -- Ajuster automatiquement la largeur des colonnes --
    for col in range(1, nb_colonnes + 1):
        lettre  = get_column_letter(col)
        max_len = 0
        for row in range(1, nb_lignes + 1):
            valeur = ws.cell(row=row, column=col).value
            if valeur:
                max_len = max(max_len, len(str(valeur)))
        # Largeur = longueur max + marge, plafonnée à 45
        ws.column_dimensions[lettre].width = min(max_len + 4, 45)

    # Hauteur fixe pour toutes les lignes de données
    for row in range(2, nb_lignes + 1):
        ws.row_dimensions[row].height = 18

    # Hauteur de l'en-tête
    ws.row_dimensions[1].height = 25


def formater_rapport(chemin_fichier):
    """Charge le fichier Excel et formate tous les onglets."""

    print(f"\n🎨 Mise en forme de : {chemin_fichier}")
    wb = load_workbook(chemin_fichier)

    for nom_onglet in wb.sheetnames:
        ws = wb[nom_onglet]
        print(f"   → Onglet '{nom_onglet}' ({ws.max_row - 1} lignes)")
        formater_onglet(ws)

    wb.save(chemin_fichier)
    print(f"   ✅ Mise en forme appliquée")